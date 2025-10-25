from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from typing import List
from sqlalchemy.ext.asyncio import AsyncSession
from DailyWeather import DailyWeather, get_calendar_month_weather


async def build_monthly_workbook(db: AsyncSession, year: int, month: int):
    """Retrieves data from database and returns a Workbook with the data for the calendar month."""

    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily Weather - {month}-{year}"

    min_temperature_font = Font(name='Verdana', size=9, color='0000FF')
    max_temperature_font = Font(name='Verdana', size=9, color='FF0000')
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # write headers
    ws.append(["Daily Weather Observations"])
    ws.append([f"{month}/{year}"])
    ws.append([])
    ws.append(["Date", "Min Temp (°C)", "Max Temp (°C)"])
    for cell in ws[ws.max_row]: # ws.max_row gives us the last row
        cell.border=thin_border

    records = await get_calendar_month_weather(db, month, year)

    row_index = ws.max_row + 1 # start after headers

    # write daily weather records
    for record in records:
        day_cell = ws.cell(row=row_index, column=1, value=record.day)
        min_temp_cell = ws.cell(row=row_index, column=2, value=record.min_temp)
        min_temp_cell.font = min_temperature_font
        max_temp_cell = ws.cell(row=row_index, column=3, value=record.max_temp)
        max_temp_cell.font = max_temperature_font
        for c in (day_cell, min_temp_cell, max_temp_cell):
            c.border = thin_border

        row_index += 1

    last_data_row = row_index - 1
    summary_start_row = row_index

    summaries = [
        ("Mean", f"=AVERAGE(B2:B{last_data_row})", f"=AVERAGE(C2:C{last_data_row})"),
        ("Min", f"=MIN(B2:B{last_data_row})", f"=MIN(C2:C{last_data_row})"),
        ("Max", f"=MAX(B2:B{last_data_row})", f"=MAX(C2:C{last_data_row})"),
    ]

    for index, (label, min_formula, max_formula) in enumerate(summaries, start=summary_start_row):
        label_cell = ws.cell(row=index, column=1, value=label)
        label_cell.number_format = "0.0"
        min_cell = ws.cell(row=index, column=2, value=min_formula)
        min_cell.number_format = "0.0"
        max_cell = ws.cell(row=index, column=3, value=max_formula)
        max_cell.number_format = "0.0"
        for c in (label_cell, min_cell, max_cell):
            c.border=thin_border
    
    return wb