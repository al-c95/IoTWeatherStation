from __future__ import annotations
import sqlite3
from pathlib import Path
from datetime import datetime, date
from typing import Optional
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent

DB_PATH = "../weather.db"
XLSX_FOLDER = "../monthly_xlsx_files"


def parse_month_anchor(value: object) -> date:
    """
    Parse the value from A3 to get the year and month.
    Accepts either a real Excel date/datetime or a string like d/mm/yyyy.
    """
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%-d/%m/%Y", "%d/%-m/%Y", "%-d/%-m/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass

        # Windows/Python often dislikes %-d / %-m, so do a fallback split.
        parts = value.split("/")
        if len(parts) == 3:
            day_s, month_s, year_s = parts
            return date(int(year_s), int(month_s), int(day_s))

    raise ValueError(f"Could not parse A3 month anchor value: {value!r}")


def parse_day(value: object) -> Optional[int]:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return int(value)

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        return int(float(value))

    raise ValueError(f"Invalid day value: {value!r}")


def parse_float(value: object) -> Optional[float]:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        return float(value)

    raise ValueError(f"Invalid numeric value: {value!r}")


def import_workbook(conn: sqlite3.Connection, xlsx_path: Path) -> int:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    anchor_date = parse_month_anchor(ws["A3"].value)
    year = anchor_date.year
    month = anchor_date.month

    inserted = 0
    row = 6

    while True:
        day_raw = ws[f"A{row}"].value
        min_raw = ws[f"B{row}"].value
        max_raw = ws[f"C{row}"].value

        # Stop if we hit a blank day cell
        if day_raw in (None, ""):
            break

        # Stop if column A is not a numeric day, e.g. "Mean"
        try:
            day = parse_day(day_raw)
        except Exception:
            break

        min_temp = parse_float(min_raw)
        max_temp = parse_float(max_raw)

        obs_date = date(year, month, day).isoformat()

        conn.execute("""
            INSERT INTO daily_weather (date, min_temp, max_temp)
            VALUES (?, ?, ?)
            ON CONFLICT(date) DO UPDATE SET
                min_temp = excluded.min_temp,
                max_temp = excluded.max_temp
        """, (obs_date, min_temp, max_temp))

        inserted += 1
        row += 1

    return inserted


def import_folder(db_path: str, folder: str) -> None:
    folder_path = Path(folder)
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    xlsx_files = sorted(
        p for p in folder_path.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    )

    if not xlsx_files:
        print("No .xlsx files found.")
        return

    conn = sqlite3.connect(db_path)
    try:
        total_rows = 0
        for path in xlsx_files:
            count = import_workbook(conn, path)
            conn.commit()
            total_rows += count
            print(f"{path.name}: imported {count} rows")

        print(f"Done. Imported/updated {total_rows} rows total.")
    finally:
        conn.close()


if __name__ == "__main__":
    import_folder(DB_PATH, XLSX_FOLDER)